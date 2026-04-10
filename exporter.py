# exporter.py - 导出 Week1 计划为 Excel

import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter
from planner import Week1Plan, DayPlan


# ─── 样式常量 ────────────────────────────────────────────────

COLOR_HEADER_BG = "1F4E79"   # 深蓝，表头背景
COLOR_HEADER_FG = "FFFFFF"   # 白色，表头文字
COLOR_DAY_BG    = "D6E4F0"   # 浅蓝，日期列背景
COLOR_CELL_ALT  = "F2F7FC"   # 极浅蓝，隔行
COLOR_GOAL_BG   = "E2EFDA"   # 浅绿，本周目标
COLOR_DONE_BG   = "FFF2CC"   # 浅黄，完成内容

THIN = Side(style="thin", color="AAAAAA")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

DAY_NAMES = ["Day1\n周一", "Day2\n周二", "Day3\n周三", "Day4\n周四",
             "Day5\n周五", "Day6\n周六", "Day7\n周日"]


def _hdr_font(size=11):
    return Font(name="微软雅黑", bold=True, color=COLOR_HEADER_FG, size=size)

def _body_font(size=10):
    return Font(name="微软雅黑", size=size)

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _wrap_align(horizontal="left", vertical="top"):
    return Alignment(wrap_text=True, horizontal=horizontal, vertical=vertical)

def _center_align():
    return Alignment(wrap_text=True, horizontal="center", vertical="center")


def _set_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:      cell.font = font
    if fill:      cell.fill = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border = border
    return cell


def _format_day_cell(dp: DayPlan) -> str:
    """将一天的任务格式化为单元格文本"""
    return (
        f"【单词】\n{dp.vocab}\n\n"
        f"【阅读】\n{dp.reading}\n\n"
        f"【{dp.other_label}】\n{dp.other}"
    )


def _add_student_sheet(wb: Workbook, plan: Week1Plan, sheet_name: str):
    ws = wb.create_sheet(title=sheet_name)

    # ── 第1行：标题 ──────────────────────────────────────────
    ws.merge_cells("A1:I1")
    title_cell = ws["A1"]
    title_cell.value = f"{plan.student_name} — {plan.exam_type} Week1 学习计划"
    title_cell.font = Font(name="微软雅黑", bold=True, size=14, color=COLOR_HEADER_BG)
    title_cell.alignment = _center_align()
    title_cell.fill = _fill("EBF3FB")
    ws.row_dimensions[1].height = 30

    # ── 第2行：列标题 ─────────────────────────────────────────
    headers = ["学员", "Day1\n周一", "Day2\n周二", "Day3\n周三", "Day4\n周四",
               "Day5\n周五", "Day6\n周六", "Day7\n周日", "本周完成内容", "本周目标"]
    for col_idx, h in enumerate(headers, start=1):
        _set_cell(
            ws, 2, col_idx, h,
            font=_hdr_font(10),
            fill=_fill(COLOR_HEADER_BG),
            alignment=_center_align(),
            border=BORDER,
        )
    ws.row_dimensions[2].height = 32

    # ── 第3行：内容 ───────────────────────────────────────────
    # 列1：学员姓名
    _set_cell(
        ws, 3, 1, plan.student_name,
        font=Font(name="微软雅黑", bold=True, size=10),
        fill=_fill(COLOR_DAY_BG),
        alignment=_center_align(),
        border=BORDER,
    )

    # 列2-8：Day1-Day7
    for i, dp in enumerate(plan.days):
        col = i + 2
        bg = COLOR_CELL_ALT if i % 2 == 0 else "FFFFFF"
        _set_cell(
            ws, 3, col, _format_day_cell(dp),
            font=_body_font(9),
            fill=_fill(bg),
            alignment=_wrap_align(),
            border=BORDER,
        )

    # 列9：本周完成内容（留空，手动填写）
    _set_cell(
        ws, 3, 9, "",
        fill=_fill(COLOR_DONE_BG),
        alignment=_wrap_align(),
        border=BORDER,
    )

    # 列10：本周目标
    _set_cell(
        ws, 3, 10, plan.week_goal,
        font=_body_font(9),
        fill=_fill(COLOR_GOAL_BG),
        alignment=_wrap_align(),
        border=BORDER,
    )

    # ── 行高 & 列宽 ───────────────────────────────────────────
    ws.row_dimensions[3].height = 160

    ws.column_dimensions["A"].width = 10   # 学员
    for col_letter in ["B", "C", "D", "E", "F", "G", "H"]:
        ws.column_dimensions[col_letter].width = 28  # Day1-7
    ws.column_dimensions["I"].width = 20   # 完成内容
    ws.column_dimensions["J"].width = 22   # 本周目标

    # 冻结前两行
    ws.freeze_panes = "A3"


def export_to_excel(plans: list) -> io.BytesIO:
    """
    将多个 Week1Plan 导出为 Excel（每人一个 sheet）。
    返回 BytesIO 对象，可直接用于 Streamlit 下载。
    """
    wb = Workbook()
    # 删除默认 sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for plan in plans:
        # sheet 名最长 31 字符，去掉非法字符
        raw_name = plan.student_name or "学员"
        safe_name = raw_name[:28].replace("/", "").replace("\\", "").replace("*", "").replace("?", "").replace("[", "").replace("]", "").replace(":", "")
        # 避免重名
        existing = wb.sheetnames
        sheet_name = safe_name
        suffix = 2
        while sheet_name in existing:
            sheet_name = f"{safe_name[:25]}_{suffix}"
            suffix += 1

        _add_student_sheet(wb, plan, sheet_name)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
